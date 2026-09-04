"""
Kasa Hareketleri ve Raporu Modülü
Logo Tiger LG_{firm}_{period}_KSLINES ve LG_{firm}_KSCARD tablolarından kasa hareketlerini listeler.
"""
import os
import io
import requests as http_requests
import pandas as pd
from sqlalchemy import text
import db_manager
from db_manager import get_engine, get_active_firm_period, get_logo_currencies, get_active_currency

BRIDGE_URL = os.environ.get('BRIDGE_URL', '').rstrip('/')
BRIDGE_KEY = os.environ.get('BRIDGE_API_KEY') or 'nexlog_bridge_2026_secure_xKj9'
USE_BRIDGE = bool(BRIDGE_URL)

def get_kasa_kartlari(conn_id=1, force_local=False):
    """Tanımlı aktif kasaları döner."""
    if not force_local and USE_BRIDGE and BRIDGE_URL:
        try:
            resp = http_requests.get(
                f"{BRIDGE_URL}/bridge/kasa/kartlar",
                headers={'X-Bridge-Key': BRIDGE_KEY, 'ngrok-skip-browser-warning': 'true'},
                timeout=15
            )
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            print(f"Bridge kasa kartlari hatası: {e}")
            return []

    try:
        fp = get_active_firm_period(conn_id)
        firm = fp.get('firm_nr', '225')
        engine = get_engine(conn_id)
        q = f"SELECT CODE, NAME FROM LG_{firm}_KSCARD WITH(NOLOCK) WHERE ACTIVE = 0 ORDER BY CODE"
        with engine.connect() as conn:
            df = pd.read_sql(text(q), conn)
            return df.to_dict(orient='records')
    except Exception as e:
        print(f"Kasa kartları getirme hatası: {e}")
        return []

def combine_invoice_and_kasa_descriptions(*parts):
    """
    Fatura arkası açıklamaları (GENEXP1..6), fatura satır açıklamaları (STLINE)
    ve kasa hareket açıklamalarını (KSLINES) akıllıca birleştirir.
    Alt dize veya birebir tekrar eden metinleri eler, en detaylı halini korur.
    """
    cleaned = []
    for part in parts:
        if not part or pd.isna(part):
            continue
        for sub in str(part).split(' - '):
            s = sub.strip()
            if not s:
                continue
            is_sub = False
            for i, existing in enumerate(cleaned):
                if s.upper() == existing.upper():
                    is_sub = True
                    break
                elif s.upper().startswith(existing.upper()) or existing.upper().startswith(s.upper()):
                    if len(s) > len(existing):
                        cleaned[i] = s
                    is_sub = True
                    break
            if not is_sub:
                cleaned.append(s)
    return ' - '.join(cleaned)

def get_kasa_hareketleri_df(filters=None, limit=2000, conn_id=1):
    """
    Filtrelere göre kasa hareketlerini DataFrame olarak döner.
    filters: dict {
        'start_date': 'YYYY-MM-DD',
        'end_date': 'YYYY-MM-DD',
        'kasa_kodu': str,
        'islem_turu': int,
        'search': str,
        'limit': int (None for all)
    }
    """
    filters = filters or {}
    fp = get_active_firm_period(conn_id)
    firm = fp.get('firm_nr', '225')
    period = fp.get('period_nr', '01')
    engine = get_engine(conn_id)

    # Logo'dan firmanın yerel para birimi bilgisini al (Örn: Cezayir Dinarı - DZD)
    logo_curr_info = get_logo_currencies(conn_id)
    local_curr = logo_curr_info.get('local_currency') or {}
    local_curcode = local_curr.get('curcode', 'DZD')
    local_curtype = local_curr.get('curtype', 81)

    where_clauses = ["KSL.CANCELLED = 0"]
    params = {
        'local_curtype': local_curtype,
        'local_curcode': local_curcode
    }

    start_date = filters.get('start_date')
    if start_date:
        where_clauses.append("CAST(KSL.DATE_ AS DATE) >= :start_date")
        params['start_date'] = start_date

    end_date = filters.get('end_date')
    if end_date:
        where_clauses.append("CAST(KSL.DATE_ AS DATE) <= :end_date")
        params['end_date'] = end_date

    kasa_kodu = filters.get('kasa_kodu')
    if kasa_kodu:
        where_clauses.append("KS.CODE = :kasa_kodu")
        params['kasa_kodu'] = kasa_kodu

    trcode = filters.get('trcode')
    if trcode is not None and str(trcode).strip() != '':
        try:
            where_clauses.append("KSL.TRCODE = :trcode")
            params['trcode'] = int(trcode)
        except ValueError:
            pass
    search = str(filters.get('search') or '').strip()
    if search:
        search_like = f"%{search}%"
        where_clauses.append(
            "(KSL.FICHENO LIKE :search OR KSL.LINEEXP LIKE :search OR "
            "KS.NAME LIKE :search OR KS.CODE LIKE :search OR TRG.GDEF LIKE :search OR TRG.GCODE LIKE :search OR "
            "CLC.CODE LIKE :search OR CLC.DEFINITION_ LIKE :search OR INV_CL.CODE LIKE :search OR INV_CL.DEFINITION_ LIKE :search OR "
            "SRV.CODE LIKE :search OR SRV.DEFINITION_ LIKE :search OR EM.CODE LIKE :search)"
        )
        params['search'] = search_like

    where_sql = " AND ".join(where_clauses)
    
    top_limit = filters.get('limit', limit)
    top_sql = f"TOP {int(top_limit)}" if top_limit and int(top_limit) > 0 else ""

    fnr_int = int(firm) if str(firm).isdigit() else 225

    sql_query = f"""
    SELECT {top_sql}
        KSL.DATE_ AS [Tarih],
        KSL.FICHENO AS [Fiş No],
        ISNULL(KSL.TRADINGGRP, '') AS [Ticari İşlem Grubu Kodu],
        ISNULL(TRG.GDEF, '') AS [Ticari İşlem Grubu Adı],
        KS.CODE AS [Kasa Kodu],
        KS.NAME AS [Kasa Adı],
        CASE KSL.TRCODE
            WHEN 1 THEN 'Nakit Tahsilat'
            WHEN 2 THEN 'Nakit Ödeme'
            WHEN 3 THEN 'Borç Dekontu'
            WHEN 4 THEN 'Alacak Dekontu'
            WHEN 11 THEN 'Cari Tahsilat'
            WHEN 12 THEN 'Cari Ödeme'
            WHEN 21 THEN 'Banka Epre/Virman'
            WHEN 22 THEN 'Bankadan Çekilen'
            WHEN 31 THEN 'Satınalma Faturası'
            WHEN 32 THEN 'Perakende Satış İade'
            WHEN 33 THEN 'Toptan Satış İade'
            WHEN 34 THEN 'Alınan Hizmet / Gider'
            WHEN 35 THEN 'Verilen Hizmet Faturası'
            WHEN 36 THEN 'Satınalma İade'
            WHEN 37 THEN 'Perakende Satış Faturası'
            WHEN 38 THEN 'Toptan Satış Faturası'
            WHEN 41 THEN 'Verilen Hizmet Faturası'
            WHEN 42 THEN 'Alınan Hizmet Faturası'
            WHEN 51 THEN 'Personel Borçlanması'
            WHEN 52 THEN 'Personele Geri Ödeme'
            WHEN 61 THEN 'Çek Tahsilatı'
            WHEN 62 THEN 'Senet Tahsilatı'
            WHEN 63 THEN 'Çek Ödemesi'
            WHEN 64 THEN 'Senet Ödemesi'
            WHEN 71 THEN 'Açılış Fişi (Borç)'
            WHEN 72 THEN 'Açılış Fişi (Alacak)'
            WHEN 73 THEN 'Kasa Virmanı (Giriş)'
            WHEN 74 THEN 'Kasa Virmanı (Çıkış)'
            WHEN 75 THEN 'Gider Pusulası'
            WHEN 76 THEN 'Müstahsil Makbuzu'
            WHEN 79 THEN 'Kur Farkı (Borç)'
            WHEN 80 THEN 'Kur Farkı (Alacak)'
            ELSE 'Diğer (' + CAST(KSL.TRCODE AS VARCHAR) + ')'
        END AS [İşlem Türü],
        -- Cari Kodu ve Ünvanı (Cari ödemeleri, tahsilatları ve faturalar)
        ISNULL(CLC.CODE, ISNULL(INV_CL.CODE, '')) AS [Cari Kodu],
        ISNULL(CLC.DEFINITION_, ISNULL(INV_CL.DEFINITION_, ISNULL(KSL.CUSTTITLE, ''))) AS [Cari / İlgili Ünvan],
        -- Gider Kodu ve Adı (Gider/Hizmet faturaları ve muhasebe kayıtları)
        ISNULL(SRV.CODE, ISNULL(EM.CODE, '')) AS [Gider Kodu],
        ISNULL(SRV.DEFINITION_, ISNULL(EM.DEFINITION_, '')) AS [Gider Adı],
        CASE KSL.SIGN
            WHEN 0 THEN KSL.AMOUNT
            ELSE 0
        END AS [Giriş (Borç)],
        CASE KSL.SIGN
            WHEN 1 THEN KSL.AMOUNT
            ELSE 0
        END AS [Çıkış (Alacak)],
        KSL.AMOUNT AS [Tutar],
        KSL.TRNET AS [Net Tutar],
        -- Logo Para Birimi (0 veya yerel tip ise DZD vb., dövizli ise USD/EUR)
        CASE 
            WHEN KSL.TRCURR = 0 OR KSL.TRCURR = :local_curtype THEN :local_curcode
            ELSE ISNULL(NULLIF(LTRIM(RTRIM(CUR.CURCODE)), ''), CAST(KSL.TRCURR AS VARCHAR))
        END AS [Para Birimi],
        -- Fatura Arkası Açıklamaları (GENEXP1..6)
        CONCAT_WS(' - ', 
            NULLIF(LTRIM(RTRIM(INV.GENEXP1)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP2)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP3)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP4)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP5)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP6)), '')
        ) AS [INV_GENEXP_ALL],
        INV_LINES.LINEEXP_ALL AS [STL_LINEEXP_ALL],
        ISNULL(KSL.LINEEXP, '') AS [KSL_LINEEXP]
    FROM LG_{firm}_{period}_KSLINES KSL WITH(NOLOCK)
    LEFT JOIN LG_{firm}_KSCARD KS WITH(NOLOCK) 
        ON KSL.CARDREF = KS.LOGICALREF
    LEFT JOIN L_TRADGRP TRG WITH(NOLOCK) 
        ON KSL.TRADINGGRP = TRG.GCODE
    -- Para Birimi Tanımı (Logo L_CURRENCYLIST)
    LEFT JOIN L_CURRENCYLIST CUR WITH(NOLOCK)
        ON (CUR.FIRMNR = {fnr_int} AND CUR.CURTYPE = KSL.TRCURR)
    -- Cari Hareket Bağlantısı (TRCODE 11, 12)
    LEFT JOIN LG_{firm}_{period}_CLFLINE CLF WITH(NOLOCK) 
        ON (KSL.TRCODE IN (11, 12) AND KSL.TRANSREF = CLF.LOGICALREF)
    LEFT JOIN LG_{firm}_CLCARD CLC WITH(NOLOCK) 
        ON CLF.CLIENTREF = CLC.LOGICALREF
    -- Fatura Bağlantısı (TRCODE 31, 34 vb.)
    LEFT JOIN LG_{firm}_{period}_INVOICE INV WITH(NOLOCK) 
        ON (KSL.TRCODE IN (31, 34, 35, 36, 37, 38, 39) AND KSL.TRANSREF = INV.LOGICALREF)
    LEFT JOIN LG_{firm}_CLCARD INV_CL WITH(NOLOCK) 
        ON INV.CLIENTREF = INV_CL.LOGICALREF
    -- Muhasebe Hesap Kartı Bağlantısı (Doğrudan muhasebe/gider hesabı)
    LEFT JOIN LG_{firm}_EMUHACC EM WITH(NOLOCK)
        ON (KSL.TRCODE NOT IN (11, 12) AND KSL.ACCREF = EM.LOGICALREF)
    -- Fatura Satırındaki Hizmet / Gider Kartı
    OUTER APPLY (
        SELECT 
            STRING_AGG(CODE, ', ') AS CODE,
            STRING_AGG(DEFINITION_, ', ') AS DEFINITION_
        FROM (
            SELECT DISTINCT SRVC.CODE, SRVC.DEFINITION_
            FROM LG_{firm}_{period}_STLINE STL WITH(NOLOCK)
            LEFT JOIN LG_{firm}_SRVCARD SRVC WITH(NOLOCK) ON STL.STOCKREF = SRVC.LOGICALREF
            WHERE STL.INVOICEREF = INV.LOGICALREF AND STL.LINETYPE = 4
        ) AS srv_distinct
    ) SRV
    -- Fatura Satırı Açıklamaları (tüm satırlar ' - ' ile birleşik)
    OUTER APPLY (
        SELECT 
            STRING_AGG(
                NULLIF(LTRIM(RTRIM(STL2.LINEEXP)), ''), 
                ' - '
            ) WITHIN GROUP (ORDER BY STL2.LOGICALREF) AS LINEEXP_ALL
        FROM LG_{firm}_{period}_STLINE STL2 WITH(NOLOCK)
        WHERE STL2.INVOICEREF = INV.LOGICALREF
          AND ISNULL(STL2.LINEEXP, '') <> ''
    ) INV_LINES
    WHERE {where_sql}
    ORDER BY KSL.DATE_ ASC, KSL.LOGICALREF ASC
    """

    with engine.connect() as conn:
        df = pd.read_sql(text(sql_query), conn, params=params)
        if not df.empty:
            df['Açıklama'] = df.apply(
                lambda r: combine_invoice_and_kasa_descriptions(
                    r.get('INV_GENEXP_ALL'),
                    r.get('STL_LINEEXP_ALL'),
                    r.get('KSL_LINEEXP')
                ), axis=1
            )
            df.drop(columns=['INV_GENEXP_ALL', 'STL_LINEEXP_ALL', 'KSL_LINEEXP'], inplace=True, errors='ignore')
        else:
            df['Açıklama'] = ''
        return df

def get_kasa_devir(filters=None, conn_id=1):
    """
    Filtrelerdeki start_date öncesindeki kasa bakiye devrini hesaplar.
    Eğer start_date verilmemişse devir 0.0 döner.
    """
    filters = filters or {}
    start_date = filters.get('start_date')
    kasa_kodu = filters.get('kasa_kodu')
    
    if not start_date:
        return 0.0
        
    fp = get_active_firm_period(conn_id)
    firm = fp.get('firm_nr', '225')
    period = fp.get('period_nr', '01')
    engine = get_engine(conn_id)
    
    where_clauses = ["KSL.CANCELLED = 0", "CAST(KSL.DATE_ AS DATE) < :start_date"]
    params = {'start_date': start_date}
    
    if kasa_kodu:
        where_clauses.append("KS.CODE = :kasa_kodu")
        params['kasa_kodu'] = kasa_kodu
        
    where_sql = " AND ".join(where_clauses)
    sql = f"""
    SELECT 
        ISNULL(SUM(CASE WHEN KSL.SIGN = 0 THEN KSL.AMOUNT ELSE -KSL.AMOUNT END), 0) AS DevirBakiye
    FROM LG_{firm}_{period}_KSLINES KSL WITH(NOLOCK)
    LEFT JOIN LG_{firm}_KSCARD KS WITH(NOLOCK) ON KSL.CARDREF = KS.LOGICALREF
    WHERE {where_sql}
    """
    try:
        with engine.connect() as conn:
            res = conn.execute(text(sql), params).fetchone()
            return float(res[0]) if res and res[0] is not None else 0.0
    except Exception as e:
        print(f"Devir bakiye hesaplama hatası: {e}")
        return 0.0

def get_kasa_data_and_summary(filters=None, limit=2000, conn_id=1, force_local=False):
    """Web arayüzü için veri listesi, devir ve kümülatif bakiye ile özet KPI toplamlarını döner."""
    if not force_local and USE_BRIDGE and BRIDGE_URL:
        try:
            payload = dict(filters or {})
            if limit: payload['limit'] = limit
            resp = http_requests.post(
                f"{BRIDGE_URL}/bridge/kasa/data-summary",
                json=payload,
                headers={'X-Bridge-Key': BRIDGE_KEY, 'ngrok-skip-browser-warning': 'true'},
                timeout=30
            )
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            print(f"Bridge get_kasa_data_and_summary hatası: {e}")

    df = get_kasa_hareketleri_df(filters, limit=limit, conn_id=conn_id)
    devir = get_kasa_devir(filters, conn_id=conn_id)

    total_giris = float(df['Giriş (Borç)'].sum()) if not df.empty else 0.0
    total_cikis = float(df['Çıkış (Alacak)'].sum()) if not df.empty else 0.0
    net_bakiye = total_giris - total_cikis
    record_count = len(df)

    # Tarihi string formata çevir (GG.AA.YYYY - 00.00.0000)
    def format_date_tr(d):
        if pd.isna(d) or d is None:
            return ''
        if hasattr(d, 'strftime'):
            return d.strftime('%d.%m.%Y')
        s = str(d).strip()
        if len(s) >= 10 and s[4] == '-' and s[7] == '-':
            return f"{s[8:10]}.{s[5:7]}.{s[0:4]}"
        return s

    if not df.empty and 'Tarih' in df.columns:
        df['Tarih_Str'] = df['Tarih'].apply(format_date_tr)
    else:
        df['Tarih_Str'] = ''

    records = df.to_dict(orient='records')
    
    # Devir ile başlayarak her satır için kümülatif BAKİYE hesapla
    running_bakiye = devir
    for r in records:
        if 'Tarih' in r:
            r['Tarih'] = format_date_tr(r['Tarih'])
        giris = float(r.get('Giriş (Borç)', 0) or 0)
        cikis = float(r.get('Çıkış (Alacak)', 0) or 0)
        running_bakiye += (giris - cikis)
        r['Bakiye'] = running_bakiye

    active_curr = get_active_currency(conn_id)

    return {
        'success': True,
        'records': records,
        'count': record_count,
        'devir': devir,
        'summary': {
            'total_giris': total_giris,
            'total_cikis': total_cikis,
            'net_bakiye': net_bakiye,
            'devir': devir,
            'son_bakiye': running_bakiye,
            'count': record_count,
            'currency': active_curr
        }
    }

def export_kasa_to_excel(filters=None, conn_id=1):
    """Filtrelenmiş kasa hareketlerini kullanıcının istediği 11 sütun ve DEVİR formatında Excel olarak üretir."""
    data_res = get_kasa_data_and_summary(filters, limit=None, conn_id=conn_id)
    records = data_res.get('records', [])
    devir = data_res.get('devir', 0.0)
    active_curr = data_res.get('summary', {}).get('currency', 'DZD')

    excel_rows = []
    
    # 1. Satır: DEVİR
    excel_rows.append({
        'KASA ADI': '',
        'TİCARİ İŞLEM GRUBU': '',
        'İŞLEM TÜRÜ': '',
        'TARİH': '',
        'FİŞ NO': '',
        'CARİ/GİDER': '',
        'DÖVİZ': active_curr,
        'AÇIKLAMA': 'DEVİR',
        'GİREN': f"{devir:,.2f} {active_curr}".replace(',', 'X').replace('.', ',').replace('X', '.') if devir >= 0 else '',
        'ÇIKAN': f"{abs(devir):,.2f} {active_curr}".replace(',', 'X').replace('.', ',').replace('X', '.') if devir < 0 else '',
        'BAKİYE': f"{devir:,.2f} {active_curr}".replace(',', 'X').replace('.', ',').replace('X', '.')
    })

    # Hareket satırları
    for r in records:
        tic_grup = str(r.get('Ticari İşlem Grubu Adı') or '').strip()
        tic_kod = str(r.get('Ticari İşlem Grubu Kodu') or '').strip()
        tic_full = f"{tic_grup} ({tic_kod})" if (tic_grup and tic_kod) else (tic_grup or tic_kod)

        cari_kodu = str(r.get('Cari Kodu') or '').strip()
        cari_unvan = str(r.get('Cari / İlgili Ünvan') or '').strip()
        gider_kodu = str(r.get('Gider Kodu') or '').strip()
        gider_adi = str(r.get('Gider Adı') or '').strip()

        # Cari/Gider birleşik
        cari_gider_parts = []
        if cari_kodu or cari_unvan:
            cari_gider_parts.append(f"{cari_kodu} / {cari_unvan}" if cari_kodu and cari_unvan else (cari_unvan or cari_kodu))
        if gider_kodu or gider_adi:
            cari_gider_parts.append(f"{gider_kodu} / {gider_adi}" if gider_kodu and gider_adi else (gider_adi or gider_kodu))
        cari_gider_str = " - ".join(cari_gider_parts) if cari_gider_parts else '-'

        row_curr = r.get('Para Birimi') or active_curr
        giris_val = float(r.get('Giriş (Borç)', 0) or 0)
        cikis_val = float(r.get('Çıkış (Alacak)', 0) or 0)
        bakiye_val = float(r.get('Bakiye', 0) or 0)

        def fmt_curr(val):
            if val == 0: return ''
            return f"{val:,.2f} {row_curr}".replace(',', 'X').replace('.', ',').replace('X', '.')

        excel_rows.append({
            'KASA ADI': r.get('Kasa Adı') or '',
            'TİCARİ İŞLEM GRUBU': tic_full,
            'İŞLEM TÜRÜ': r.get('İşlem Türü') or '',
            'TARİH': r.get('Tarih') or '',
            'FİŞ NO': str(r.get('Fiş No') or '').lstrip('0') or r.get('Fiş No') or '',
            'CARİ/GİDER': cari_gider_str,
            'DÖVİZ': row_curr,
            'AÇIKLAMA': r.get('Açıklama') or '',
            'GİREN': fmt_curr(giris_val),
            'ÇIKAN': fmt_curr(cikis_val),
            'BAKİYE': f"{bakiye_val:,.2f} {row_curr}".replace(',', 'X').replace('.', ',').replace('X', '.')
        })

    excel_df = pd.DataFrame(excel_rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        excel_df.to_excel(writer, index=False, sheet_name='Kasa_Hareketleri')
    output.seek(0)
    return output

def get_kasa_ozet_raporu(filters=None, conn_id=1, force_local=False):
    """
    Her kasanın KODU, KASA ADI, GİREN, ÇIKAN, BAKİYE ve PARA BİRİMİ özetini döner.
    filters: {'start_date': '...', 'end_date': '...', 'search': '...'}
    """
    if not force_local and USE_BRIDGE and BRIDGE_URL:
        try:
            resp = http_requests.post(
                f"{BRIDGE_URL}/bridge/kasa/ozet-raporu",
                json=filters or {},
                headers={'X-Bridge-Key': BRIDGE_KEY, 'ngrok-skip-browser-warning': 'true'},
                timeout=30
            )
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            print(f"Bridge get_kasa_ozet_raporu hatası: {e}")

    filters = filters or {}
    fp = get_active_firm_period(conn_id)
    firm = fp.get('firm_nr', '225')
    period = fp.get('period_nr', '01')
    engine = get_engine(conn_id)
    active_curr = get_active_currency(conn_id)

    where_kslines = ["KSL.CANCELLED = 0"]
    params = {}

    start_date = filters.get('start_date')
    if start_date:
        where_kslines.append("CAST(KSL.DATE_ AS DATE) >= :start_date")
        params['start_date'] = start_date

    end_date = filters.get('end_date')
    if end_date:
        where_kslines.append("CAST(KSL.DATE_ AS DATE) <= :end_date")
        params['end_date'] = end_date

    kslines_where = " AND ".join(where_kslines)

    where_ks = ["KS.ACTIVE = 0"]
    search = str(filters.get('search') or '').strip()
    if search:
        search_like = f"%{search}%"
        where_ks.append("(KS.CODE LIKE :search OR KS.NAME LIKE :search)")
        params['search'] = search_like

    ks_where = " AND ".join(where_ks)

    sql_query = f"""
    SELECT 
        KS.CODE AS [KODU],
        KS.NAME AS [KASA ADI],
        ISNULL(SUM(CASE WHEN KSL.SIGN = 0 THEN KSL.AMOUNT ELSE 0 END), 0) AS [GİREN],
        ISNULL(SUM(CASE WHEN KSL.SIGN = 1 THEN KSL.AMOUNT ELSE 0 END), 0) AS [ÇIKAN],
        ISNULL(SUM(CASE WHEN KSL.SIGN = 0 THEN KSL.AMOUNT ELSE -KSL.AMOUNT END), 0) AS [BAKİYE],
        '{active_curr}' AS [PARA BİRİMİ]
    FROM LG_{firm}_KSCARD KS WITH(NOLOCK)
    LEFT JOIN LG_{firm}_{period}_KSLINES KSL WITH(NOLOCK) 
        ON KSL.CARDREF = KS.LOGICALREF AND {kslines_where}
    WHERE {ks_where}
    GROUP BY KS.CODE, KS.NAME
    ORDER BY KS.CODE
    """

    with engine.connect() as conn:
        df = pd.read_sql(text(sql_query), conn, params=params)

    total_giren = float(df['GİREN'].sum()) if not df.empty else 0.0
    total_cikan = float(df['ÇIKAN'].sum()) if not df.empty else 0.0
    total_bakiye = float(df['BAKİYE'].sum()) if not df.empty else 0.0
    negative_count = int((df['BAKİYE'] < 0).sum()) if not df.empty else 0

    records = df.to_dict(orient='records')
    return {
        'success': True,
        'records': records,
        'count': len(records),
        'summary': {
            'total_giren': total_giren,
            'total_cikan': total_cikan,
            'total_bakiye': total_bakiye,
            'negative_count': negative_count,
            'kasa_count': len(records),
            'currency': active_curr
        }
    }

def export_kasa_raporu_to_excel(filters=None, conn_id=1):
    """Kasa bakiye raporunu Excel (.xlsx) dosyası olarak üretir."""
    res = get_kasa_ozet_raporu(filters, conn_id=conn_id)
    df = pd.DataFrame(res['records'])
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Kasa_Bakiye_Raporu')
    output.seek(0)
    return output

def get_kasa_ticari_grup_analizi(filters=None, conn_id=1, force_local=False):
    """
    Ticari İşlem Grubu bazında 12 aylık Kasa Analiz pivot tablosunu döner.
    filters: dict {
        'year': int (varsayılan 2026),
        'kasa_kodu': str (opsiyonel),
        'direction': 'all' | 'cikis' | 'giris' (varsayılan 'all'),
        'include_empty': bool (varsayılan True)
    }
    """
    if not force_local and USE_BRIDGE and BRIDGE_URL:
        try:
            resp = http_requests.post(
                f"{BRIDGE_URL}/bridge/kasa/analiz-ticari-grup",
                json=filters or {},
                headers={'X-Bridge-Key': BRIDGE_KEY, 'ngrok-skip-browser-warning': 'true'},
                timeout=45
            )
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            print(f"Bridge get_kasa_ticari_grup_analizi hatası: {e}")

    filters = filters or {}
    fp = get_active_firm_period(conn_id)
    firm = fp.get('firm_nr', '225')
    period = fp.get('period_nr', '01')
    engine = get_engine(conn_id)
    active_curr = get_active_currency(conn_id) or 'DZD'

    year = filters.get('year')
    try:
        year = int(year) if year else 2026
    except (ValueError, TypeError):
        year = 2026

    kasa_kodu = filters.get('kasa_kodu')
    direction = (filters.get('direction') or 'all').lower()
    include_empty = filters.get('include_empty', True)
    if isinstance(include_empty, str):
        include_empty = include_empty.lower() in ['true', '1', 'yes']

    where_clauses = ["YEAR(KSL.DATE_) = :year"]
    params = {'year': year}

    if kasa_kodu and kasa_kodu != 'ALL' and str(kasa_kodu).strip():
        where_clauses.append("KS.CODE = :kasa_kodu")
        params['kasa_kodu'] = str(kasa_kodu).strip()

    if direction == 'cikis':
        where_clauses.append("KSL.SIGN = 1")
    elif direction == 'giris':
        where_clauses.append("KSL.SIGN = 0")

    if not include_empty:
        where_clauses.append("ISNULL(LTRIM(RTRIM(KSL.TRADINGGRP)), '') <> ''")

    where_sql = " AND ".join(where_clauses)

    sql_query = f"""
    SELECT 
        COALESCE(NULLIF(LTRIM(RTRIM(TRG.GDEF)), ''), NULLIF(LTRIM(RTRIM(KSL.TRADINGGRP)), ''), '(Grup Belirtilmemiş)') AS [TICARI_ISLEM_GRUBU],
        ISNULL(NULLIF(LTRIM(RTRIM(KSL.TRADINGGRP)), ''), '-') AS [GRUP_KODU],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 1 THEN KSL.AMOUNT ELSE 0 END), 0) AS [OCAK],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 2 THEN KSL.AMOUNT ELSE 0 END), 0) AS [SUBAT],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 3 THEN KSL.AMOUNT ELSE 0 END), 0) AS [MART],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 4 THEN KSL.AMOUNT ELSE 0 END), 0) AS [NISAN],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 5 THEN KSL.AMOUNT ELSE 0 END), 0) AS [MAYIS],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 6 THEN KSL.AMOUNT ELSE 0 END), 0) AS [HAZIRAN],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 7 THEN KSL.AMOUNT ELSE 0 END), 0) AS [TEMMUZ],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 8 THEN KSL.AMOUNT ELSE 0 END), 0) AS [AGUSTOS],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 9 THEN KSL.AMOUNT ELSE 0 END), 0) AS [EYLUL],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 10 THEN KSL.AMOUNT ELSE 0 END), 0) AS [EKIM],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 11 THEN KSL.AMOUNT ELSE 0 END), 0) AS [KASIM],
        ISNULL(SUM(CASE WHEN MONTH(KSL.DATE_) = 12 THEN KSL.AMOUNT ELSE 0 END), 0) AS [ARALIK],
        ISNULL(SUM(KSL.AMOUNT), 0) AS [TOPLAM]
    FROM LG_{firm}_{period}_KSLINES KSL WITH(NOLOCK)
    LEFT JOIN LG_{firm}_KSCARD KS WITH(NOLOCK) 
        ON KSL.CARDREF = KS.LOGICALREF
    LEFT JOIN L_TRADGRP TRG WITH(NOLOCK) 
        ON KSL.TRADINGGRP = TRG.GCODE
    WHERE {where_sql}
    GROUP BY TRG.GDEF, KSL.TRADINGGRP
    ORDER BY [TOPLAM] DESC
    """

    with engine.connect() as conn:
        df = pd.read_sql(text(sql_query), conn, params=params)

    months = ['OCAK', 'SUBAT', 'MART', 'NISAN', 'MAYIS', 'HAZIRAN', 'TEMMUZ', 'AGUSTOS', 'EYLUL', 'EKIM', 'KASIM', 'ARALIK']
    monthly_totals = {m: float(df[m].sum()) if not df.empty else 0.0 for m in months}
    grand_total = float(df['TOPLAM'].sum()) if not df.empty else 0.0

    # Kategorize Etme Mantığı
    def categorize_trading_group(gcode, gdef):
        name = str(gdef or '').upper()
        if any(k in name for k in ['AKARYAKIT', 'ARAÇ', 'ARAC', 'NAKLİYE', 'NAKLIYE', 'SERVİS', 'SERVIS', 'YEDEK PARÇA', 'YEDEK PARCA', 'MUAYENE', 'SİGORTA', 'SIGORTA', 'HOWO', 'KAMYON', 'BİNEK', 'BINEK']):
            return ('ARAÇ, ULAŞIM & AKARYAKIT GİDERLERİ', 'ARAÇ & ULAŞIM GİDERLERİ TOPLAMI', 1)
        if any(k in name for k in ['AVANS']):
            return ('PERSONEL AVANS GİDERLERİ', 'PERSONEL AVANS TOPLAMI', 3)
        if any(k in name for k in ['MAAŞ', 'MAAS', 'MESAİ', 'MESAI', 'ÜCRET', 'UCRET', 'İKRAMİYE', 'IKRAMIYE', 'İZİN', 'IZIN', 'ÇALIŞMA İZNİ', 'OTURUM VERGİ', 'VİZE ÜCRET']):
            return ('PERSONEL MAAŞ & ÖZLÜK GİDERLERİ', 'PERSONEL MAAŞ & ÖZLÜK TOPLAMI', 2)
        if any(k in name for k in ['MALZEME', 'HURDA', 'SATICILAR', 'ALIMLARI']):
            return ('OPERASYONEL SATINALMA & HURDA GİDERLERİ', 'SATINALMA & HURDA TOPLAMI', 4)
        if any(k in name for k in ['MUTFAK', 'İAŞE', 'IASE', 'YEMEK', 'SU GİDER', 'TÜP GAZ', 'TUP GAZ', 'KİRA', 'KIRA', 'KIRTASİYE', 'OFİS', 'OFIS', 'İLETİŞİM', 'TELEFON', 'İNTERNET', 'KARGO', 'SEYAHAT', 'GÜMRÜK', 'GUMRUK', 'SAĞLIK', 'SAGLIK', 'DANIŞMANLIK', 'VERGİ', 'VERGI', 'TAMİR', 'TAMIR', 'BAĞIŞ', 'BAGIS']):
            return ('İDARİ OFİS, TESİS & GENEL GİDERLER', 'İDARİ & TESİS GİDERLERİ TOPLAMI', 5)
        return ('DİĞER OPERASYONEL GİDERLER', 'DİĞER GİDERLER TOPLAMI', 6)

    records = df.to_dict(orient='records')
    category_map = {}

    for r in records:
        cat_name, cat_lbl, cat_order = categorize_trading_group(r.get('GRUP_KODU'), r.get('TICARI_ISLEM_GRUBU'))
        r['KATEGORI'] = cat_name
        r['KATEGORI_LABEL'] = cat_lbl
        r['KATEGORI_ORDER'] = cat_order

        if cat_name not in category_map:
            category_map[cat_name] = {
                'name': cat_name,
                'total_label': cat_lbl,
                'order': cat_order,
                'records': [],
                'monthly_totals': {m: 0.0 for m in months},
                'grand_total': 0.0
            }

        category_map[cat_name]['records'].append(r)
        category_map[cat_name]['grand_total'] += float(r.get('TOPLAM', 0.0))
        for m in months:
            category_map[cat_name]['monthly_totals'][m] += float(r.get(m, 0.0))

    categories = list(category_map.values())
    categories.sort(key=lambda x: x['order'])

    return {
        'success': True,
        'records': records,
        'categories': categories,
        'monthly_totals': monthly_totals,
        'grand_total': grand_total,
        'count': len(records),
        'currency': active_curr,
        'year': year
    }

def export_kasa_analizi_to_excel(filters=None, conn_id=1):
    """
    Ticari İşlem Grubu analizi Excel dosyasını (Kullanıcının talep ettiği kategori bloklu pivot formatında) üretir.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    res = get_kasa_ticari_grup_analizi(filters, conn_id=conn_id)
    categories = res.get('categories', [])
    monthly_totals = res.get('monthly_totals', {})
    grand_total = res.get('grand_total', 0.0)
    year = res.get('year', 2026)
    curr = res.get('currency', 'DZD')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Kasa_Analizi_{year}"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name='Segoe UI', size=10, bold=True, color='00E5FF')
    header_fill = PatternFill(start_color='081020', end_color='081020', fill_type='solid')
    thin_border_side = Side(style='thin', color='CBD5E1')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    num_font = Font(name='Consolas', size=10, bold=True, color='000000')
    grp_font = Font(name='Segoe UI', size=10, bold=True, color='000000')
    month_keys = ['OCAK', 'SUBAT', 'MART', 'NISAN', 'MAYIS', 'HAZIRAN', 'TEMMUZ', 'AGUSTOS', 'EYLUL', 'EKIM', 'KASIM', 'ARALIK']

    row_idx = 1

    for cat in categories:
        # Kategori Başlık Satırı
        headers = [cat['name'], 'OCAK', 'ŞUBAT', 'MART', 'NİSAN', 'MAYIS', 'HAZİRAN', 'TEMMUZ', 'AĞUSTOS', 'EYLÜL', 'EKİM', 'KASIM', 'ARALIK', 'TOPLAM']
        ws.row_dimensions[row_idx].height = 26
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = cell_border
            cell.alignment = Alignment(horizontal='center' if col_idx == 1 else 'right', vertical='center')
        row_idx += 1

        # Kategori Veri Satırları
        for r in cat['records']:
            ws.row_dimensions[row_idx].height = 22
            cell_grp = ws.cell(row=row_idx, column=1, value=r.get('TICARI_ISLEM_GRUBU'))
            cell_grp.font = grp_font
            cell_grp.border = cell_border
            cell_grp.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for m_idx, m_key in enumerate(month_keys, 2):
                val = float(r.get(m_key, 0.0))
                cell_m = ws.cell(row=row_idx, column=m_idx, value=val)
                cell_m.font = num_font
                cell_m.number_format = '#,##0.00'
                cell_m.border = cell_border
                cell_m.alignment = Alignment(horizontal='right', vertical='center')

            tot_val = float(r.get('TOPLAM', 0.0))
            cell_tot = ws.cell(row=row_idx, column=14, value=tot_val)
            cell_tot.font = Font(name='Consolas', size=10, bold=True, color='000000')
            cell_tot.number_format = '#,##0.00'
            cell_tot.border = cell_border
            cell_tot.alignment = Alignment(horizontal='right', vertical='center')
            row_idx += 1

        # Kategori Alt Toplam Satırı (Buz Mavisi Zemin, Canlı Camgöbeği Üst Çizgi)
        ws.row_dimensions[row_idx].height = 24
        subtotal_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
        subtotal_border = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=Side(style='medium', color='00D2FF'),
            bottom=thin_border_side
        )

        cell_cat_tot = ws.cell(row=row_idx, column=1, value=cat['total_label'])
        cell_cat_tot.font = Font(name='Segoe UI', size=10, bold=True, color='000000')
        cell_cat_tot.fill = subtotal_fill
        cell_cat_tot.border = subtotal_border
        cell_cat_tot.alignment = Alignment(horizontal='center', vertical='center')

        for m_idx, m_key in enumerate(month_keys, 2):
            m_val = float(cat['monthly_totals'].get(m_key, 0.0))
            c_m = ws.cell(row=row_idx, column=m_idx, value=m_val)
            c_m.font = Font(name='Consolas', size=10, bold=True, color='000000')
            c_m.number_format = '#,##0.00'
            c_m.fill = subtotal_fill
            c_m.border = subtotal_border
            c_m.alignment = Alignment(horizontal='right', vertical='center')

        c_grand_cat = ws.cell(row=row_idx, column=14, value=cat['grand_total'])
        c_grand_cat.font = Font(name='Consolas', size=10, bold=True, color='000000')
        c_grand_cat.number_format = '#,##0.00'
        c_grand_cat.fill = subtotal_fill
        c_grand_cat.border = subtotal_border
        c_grand_cat.alignment = Alignment(horizontal='right', vertical='center')
        row_idx += 2 # Bir satır boşluk bırak

    # GENEL TOPLAM Satırı (En Alt)
    ws.row_dimensions[row_idx].height = 28
    grand_fill = PatternFill(start_color='081020', end_color='081020', fill_type='solid')
    grand_border = Border(
        left=thin_border_side, 
        right=thin_border_side, 
        top=Side(style='medium', color='00E5FF'), 
        bottom=Side(style='double', color='00E5FF')
    )

    cell_tot_lbl = ws.cell(row=row_idx, column=1, value="GENEL GİDERLER TOPLAMI")
    cell_tot_lbl.font = Font(name='Segoe UI', size=11, bold=True, color='00E5FF')
    cell_tot_lbl.fill = grand_fill
    cell_tot_lbl.border = grand_border
    cell_tot_lbl.alignment = Alignment(horizontal='center', vertical='center')

    for m_idx, m_key in enumerate(month_keys, 2):
        m_tot = float(monthly_totals.get(m_key, 0.0))
        cell_m_tot = ws.cell(row=row_idx, column=m_idx, value=m_tot)
        cell_m_tot.font = Font(name='Consolas', size=10, bold=True, color='00E5FF')
        cell_m_tot.number_format = '#,##0.00'
        cell_m_tot.fill = grand_fill
        cell_m_tot.border = grand_border
        cell_m_tot.alignment = Alignment(horizontal='right', vertical='center')

    cell_grand = ws.cell(row=row_idx, column=14, value=grand_total)
    cell_grand.font = Font(name='Consolas', size=11, bold=True, color='34D399')
    cell_grand.number_format = '#,##0.00'
    cell_grand.fill = grand_fill
    cell_grand.border = grand_border
    cell_grand.alignment = Alignment(horizontal='right', vertical='center')

    # Sütun Genişlikleri
    ws.column_dimensions['A'].width = 38
    for col_let in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col_let].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def get_kasa_analiz_drilldown(filters=None, conn_id=1, force_local=False):
    """
    Kasa Analiz tablosunda tıklanan ay ve ticari işlem grubuna ait detay kasa hareket satırlarını (fişleri) getirir.
    all_groups=True ise o ayın TÜM gruplarına ait satırları getirir (aylık toplam hücresine tıklanınca).
    """
    if not force_local and USE_BRIDGE and BRIDGE_URL:
        try:
            resp = http_requests.post(
                f"{BRIDGE_URL}/bridge/kasa/analiz-drilldown",
                json=filters or {},
                headers={'X-Bridge-Key': BRIDGE_KEY, 'ngrok-skip-browser-warning': 'true'},
                timeout=30
            )
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            print(f"Bridge get_kasa_analiz_drilldown hatası: {e}")

    if filters is None:
        filters = {}

    fp = get_active_firm_period(conn_id)
    firm = fp.get('firm_nr', '225')
    period = fp.get('period_nr', '01')
    local_curcode = get_active_currency(conn_id) or 'DZD'
    local_curtype = 0
    fnr_int = int(firm) if str(firm).isdigit() else 225

    trading_grp = str(filters.get('trading_grp') or '').strip()
    trading_code = str(filters.get('trading_code') or '').strip()
    year = int(filters.get('year') or 2026)
    month = int(filters.get('month') or 1)
    kasa_kodu = str(filters.get('kasa_kodu') or 'ALL').strip()

    direction = str(filters.get('direction') or 'cikis').strip()
    all_groups = bool(filters.get('all_groups', False))

    # GRUP_KODU '-' ise boş grup olarak işle
    if trading_code == '-':
        trading_code = ''

    where_clauses = [
        "YEAR(KSL.DATE_) = :year",
        "MONTH(KSL.DATE_) = :month"
    ]
    params = {
        'year': year,
        'month': month,
    }

    if direction == 'cikis':
        where_clauses.append("KSL.SIGN = 1")
    elif direction == 'giris':
        where_clauses.append("KSL.SIGN = 0")

    if kasa_kodu and kasa_kodu != 'ALL':
        where_clauses.append("KS.CODE = :kasa_kodu")
        params['kasa_kodu'] = kasa_kodu

    # Grup eşleşmesi — all_groups=True ise grup filtresi uygulanmaz (tüm gruplar döner)
    if not all_groups:
        if trading_code and trading_code not in ('', '-'):
            where_clauses.append("ISNULL(LTRIM(RTRIM(KSL.TRADINGGRP)), '') = :trading_code")
            params['trading_code'] = trading_code.strip()
        elif trading_grp in ['(Grup Belirtilmemiş)', '(Grup Belirtilmem)', 'Grup Belirtilmemiş', '']:
            where_clauses.append("(KSL.TRADINGGRP IS NULL OR LTRIM(RTRIM(KSL.TRADINGGRP)) = '')")
        else:
            where_clauses.append("(TRG.GDEF = :trading_grp OR ISNULL(LTRIM(RTRIM(KSL.TRADINGGRP)), '') = :trading_grp)")
            params['trading_grp'] = trading_grp.strip()

    where_sql = " AND ".join(where_clauses)
    print(f"[DRILLDOWN] year={year} month={month} grp='{trading_grp}' code='{trading_code}' all_groups={all_groups} direction={direction} WHERE: {where_sql} PARAMS: {params}")

    sql_query = f"""
    SELECT 
        KSL.LOGICALREF,
        CONVERT(VARCHAR(10), KSL.DATE_, 104) AS [Tarih],
        ISNULL(KSL.FICHENO, '') AS [Fiş No],
        KS.CODE AS [Kasa Kodu],
        KS.NAME AS [Kasa Adı],
        ISNULL(TRG.GDEF, ISNULL(KSL.TRADINGGRP, '(Grup Belirtilmemiş)')) AS [Grup Adı],
        CASE KSL.TRCODE
            WHEN 1 THEN 'Nakit Tahsilat'
            WHEN 2 THEN 'Nakit Ödeme'
            WHEN 3 THEN 'Borç Dekontu'
            WHEN 4 THEN 'Alacak Dekontu'
            WHEN 11 THEN 'Cari Tahsilat'
            WHEN 12 THEN 'Cari Ödeme'
            WHEN 21 THEN 'Banka Virman'
            WHEN 22 THEN 'Bankadan Çekilen'
            WHEN 31 THEN 'Satınalma Faturası'
            WHEN 34 THEN 'Alınan Hizmet / Gider'
            WHEN 35 THEN 'Verilen Hizmet Faturası'
            WHEN 51 THEN 'Personel Borçlanması'
            WHEN 52 THEN 'Personele Geri Ödeme'
            WHEN 71 THEN 'Açılış Fişi'
            WHEN 73 THEN 'Kasa Virmanı (Giriş)'
            WHEN 74 THEN 'Kasa Virmanı (Çıkış)'
            ELSE 'İşlem (' + CAST(KSL.TRCODE AS VARCHAR) + ')'
        END AS [İşlem Türü],
        ISNULL(CLC.CODE, ISNULL(INV_CL.CODE, '')) AS [Cari Kodu],
        ISNULL(CLC.DEFINITION_, ISNULL(INV_CL.DEFINITION_, ISNULL(KSL.CUSTTITLE, ''))) AS [Cari Ünvanı],
        -- Fatura Arkası Açıklamaları (GENEXP1..6)
        CONCAT_WS(' - ', 
            NULLIF(LTRIM(RTRIM(INV.GENEXP1)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP2)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP3)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP4)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP5)), ''),
            NULLIF(LTRIM(RTRIM(INV.GENEXP6)), '')
        ) AS [INV_GENEXP_ALL],
        INV_LINES.LINEEXP_ALL AS [STL_LINEEXP_ALL],
        ISNULL(KSL.LINEEXP, '') AS [KSL_LINEEXP],
        KSL.AMOUNT AS [Tutar]
    FROM LG_{firm}_{period}_KSLINES KSL WITH(NOLOCK)
    LEFT JOIN LG_{firm}_KSCARD KS WITH(NOLOCK) ON KSL.CARDREF = KS.LOGICALREF
    LEFT JOIN L_TRADGRP TRG WITH(NOLOCK) ON KSL.TRADINGGRP = TRG.GCODE
    LEFT JOIN LG_{firm}_{period}_CLFLINE CLF WITH(NOLOCK) ON (KSL.TRCODE IN (11, 12) AND KSL.TRANSREF = CLF.LOGICALREF)
    LEFT JOIN LG_{firm}_CLCARD CLC WITH(NOLOCK) ON CLF.CLIENTREF = CLC.LOGICALREF
    LEFT JOIN LG_{firm}_{period}_INVOICE INV WITH(NOLOCK) ON (KSL.TRCODE IN (31, 34, 35, 36, 37, 38, 39) AND KSL.TRANSREF = INV.LOGICALREF)
    LEFT JOIN LG_{firm}_CLCARD INV_CL WITH(NOLOCK) ON INV.CLIENTREF = INV_CL.LOGICALREF
    OUTER APPLY (
        SELECT 
            STRING_AGG(
                NULLIF(LTRIM(RTRIM(STL2.LINEEXP)), ''), 
                ' - '
            ) WITHIN GROUP (ORDER BY STL2.LOGICALREF) AS LINEEXP_ALL
        FROM LG_{firm}_{period}_STLINE STL2 WITH(NOLOCK)
        WHERE STL2.INVOICEREF = INV.LOGICALREF
          AND ISNULL(STL2.LINEEXP, '') <> ''
    ) INV_LINES
    WHERE {where_sql}
    ORDER BY KSL.DATE_ ASC, KSL.AMOUNT DESC
    """

    engine = get_engine(conn_id)
    with engine.connect() as conn:
        df = pd.read_sql(text(sql_query), conn, params=params)

    if not df.empty:
        df['Açıklama'] = df.apply(
            lambda r: combine_invoice_and_kasa_descriptions(
                r.get('INV_GENEXP_ALL'),
                r.get('STL_LINEEXP_ALL'),
                r.get('KSL_LINEEXP')
            ), axis=1
        )
        df.drop(columns=['INV_GENEXP_ALL', 'STL_LINEEXP_ALL', 'KSL_LINEEXP'], inplace=True, errors='ignore')
    else:
        df['Açıklama'] = ''

    lines = df.to_dict(orient='records')
    total_amount = float(df['Tutar'].sum()) if not df.empty else 0.0

    return {
        'success': True,
        'lines': lines,
        'count': len(lines),
        'total': total_amount,
        'currency': local_curcode,
        'trading_grp': trading_grp,
        'year': year,
        'month': month
    }
