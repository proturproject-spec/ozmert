import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _format_currency_val(val):
    try:
        return float(val or 0.0)
    except (ValueError, TypeError):
        return 0.0

def _get_doviz_str(dov):
    dov = str(dov or '').strip().upper()
    if dov in ('0', 'TL', 'TRY', ''):
        return 'TL'
    elif dov in ('1', 'USD', '$'):
        return 'USD'
    elif dov in ('20', 'EUR', '€'):
        return 'EUR'
    return dov

def generate_eski_sablon_excel(
    selected_date,
    bank_records,
    incoming_transfers,
    customer_checks,
    outgoing_transfers,
    own_checks,
    credits,
    credit_cards,
    custom_payments,
    next_day_payments=None,
    next_business_date_formatted="",
    total_bank_balance_tl=0.0,
    total_bank_balance_usd=0.0,
    total_bank_balance_eur=0.0
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Günlük Nakit Akış"
    ws.views.sheetView[0].showGridLines = True

    # Palette
    font_family = "Segoe UI"
    
    # Styles
    title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    section_green_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    section_red_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    section_blue_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    section_amber_fill = PatternFill(start_color="92400E", end_color="92400E", fill_type="solid")
    section_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")

    subhdr_tah_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    subhdr_tah_font = Font(name=font_family, size=10, bold=True, color="065F46")
    
    subhdr_ode_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    subhdr_ode_font = Font(name=font_family, size=10, bold=True, color="991B1B")

    th_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    th_font = Font(name=font_family, size=9, bold=True, color="334155")
    
    total_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    total_font = Font(name=font_family, size=9, bold=True, color="0F172A")
    
    data_font = Font(name=font_family, size=9, color="1E293B")
    data_sub_font = Font(name=font_family, size=8, color="64748B", italic=True)

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    thick_top_border = Border(top=Side(border_style="medium", color="475569"), bottom=Side(border_style="double", color="475569"), left=thin_border_side, right=thin_border_side)

    # Number Formats
    num_fmt_tl = '#,##0.00 "₺"'
    num_fmt_usd = '#,##0.00 "$"'
    num_fmt_eur = '#,##0.00 "€"'

    current_row = 1

    # --- TITLE HEADER ---
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
    title_cell = ws.cell(row=current_row, column=1, value=f"GÜNLÜK NAKİT AKIŞ RAPORU — {selected_date}")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 32
    current_row += 2

    # --- MAIN TWO COLUMNS: LEFT (TAHSİLATLAR), RIGHT (ÖDEMELER) ---
    # Left columns: A (1) to E (5) -> Hesap/Kasa, Açıklama, TL, USD, EUR
    # Gap column: F (6)
    # Right columns: G (7) to K (11) -> Hesap/Kasa, Açıklama, TL, USD, EUR

    # Header row for sections
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    c_left = ws.cell(row=current_row, column=1, value="TAHSİLATLAR (GİRİŞLER)")
    c_left.font = section_font
    c_left.fill = section_green_fill
    c_left.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=11)
    c_right = ws.cell(row=current_row, column=7, value="ÖDEMELER (ÇIKIŞLAR)")
    c_right.font = section_font
    c_right.fill = section_red_fill
    c_right.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # Helper function to write table headers
    def write_headers(row, col_start):
        headers = ["FİNANSAL HESAP / KASA", "AÇIKLAMA", "TL", "USD", "EUR"]
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=col_start + i, value=h)
            c.font = th_font
            c.fill = th_fill
            c.border = cell_border
            if i >= 2:
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20

    # Build Left Column Data (Tahsilatlar)
    # 1. Banka / Çek Tahsilatları
    left_rows = [] # list of tuples: (hesap, aciklama, tl, usd, eur)
    left_sections = [] # list of dicts: {'title': str, 'rows': list}
    
    sec1_rows = []
    for t in (incoming_transfers or []):
        hturu = str(t.get('HESAP_TURU_RAPOR') or '')
        if hturu != 'Kasa':
            dov = _get_doviz_str(t.get('HESAP_DOVIZI_RAPOR'))
            val = _format_currency_val(t.get('TL_TUTAR') if dov == 'TL' else t.get('DOVIZLI_TUTAR'))
            hesap = str(t.get('HESAP_ACIKLAMASI') or '-')
            acik = str(t.get('CARI_UNVAN') or t.get('SATIR_ACIKLAMASI') or t.get('FIS_ACIKLAMASI') or '-')
            sec1_rows.append((hesap, acik, val if dov == 'TL' else 0.0, val if dov == 'USD' else 0.0, val if dov == 'EUR' else 0.0))

    for c in (customer_checks or []):
        if hasattr(c, 'to_dict'):
            c = c.to_dict()
        val = _format_currency_val(c.get('TUTAR'))
        dov_code = str(c.get('DOVIZ_TIPI') or '0')
        dov = _get_doviz_str(dov_code)
        hesap = str(c.get('CH_UNVANI') or c.get('CARI_UNVAN') or 'Müşteri Çeki')
        acik = f"Müşteri Çeki - {c.get('BANKA') or ''} (Vade: {str(c.get('VADE') or '')[:10]})"
        sec1_rows.append((hesap, acik, val if dov == 'TL' else 0.0, val if dov == 'USD' else 0.0, val if dov == 'EUR' else 0.0))

    left_sections.append({'title': "1. BANKA / ÇEK TAHSİLATLARI", 'rows': sec1_rows})

    # 2. Kasa Tahsilatları
    sec2_rows = []
    for t in (incoming_transfers or []):
        hturu = str(t.get('HESAP_TURU_RAPOR') or '')
        if hturu == 'Kasa':
            dov = _get_doviz_str(t.get('HESAP_DOVIZI_RAPOR'))
            val = _format_currency_val(t.get('TL_TUTAR') if dov == 'TL' else t.get('DOVIZLI_TUTAR'))
            hesap = str(t.get('HESAP_ACIKLAMASI') or '-')
            acik = str(t.get('CARI_UNVAN') or t.get('SATIR_ACIKLAMASI') or t.get('FIS_ACIKLAMASI') or '-')
            sec2_rows.append((hesap, acik, val if dov == 'TL' else 0.0, val if dov == 'USD' else 0.0, val if dov == 'EUR' else 0.0))

    left_sections.append({'title': "2. KASA TAHSİLATLARI", 'rows': sec2_rows})

    # Build Right Column Data (Ödemeler)
    # 1. Banka Ödemeleri
    right_sections = []
    sec_ode1_rows = []
    sec_cc_rows = []
    sec_kasa_rows = []

    for t in (outgoing_transfers or []):
        hkodu = str(t.get('HESAP_KODU') or '')
        hturu = str(t.get('HESAP_TURU_RAPOR') or '')
        fturu = str(t.get('FIS_TURU') or '')
        satir = str(t.get('SATIR_ACIKLAMASI') or '')
        fis = str(t.get('FIS_ACIKLAMASI') or '')
        
        is_cc = hkodu.startswith('50.') or 'Kredi Kart' in hturu or 'Kredi Kart' in fturu or 'Kredi Kart' in satir or 'Kredi Kart' in fis
        dov = _get_doviz_str(t.get('HESAP_DOVIZI_RAPOR'))
        val = _format_currency_val(t.get('TL_TUTAR') if dov == 'TL' else t.get('DOVIZLI_TUTAR'))
        hesap = str(t.get('HESAP_ACIKLAMASI') or '-')
        acik = str(t.get('CARI_UNVAN') or satir or fis or '-')
        
        item = (hesap, acik, val if dov == 'TL' else 0.0, val if dov == 'USD' else 0.0, val if dov == 'EUR' else 0.0)
        
        if hturu == 'Kasa':
            sec_kasa_rows.append(item)
        elif is_cc:
            sec_cc_rows.append(item)
        else:
            sec_ode1_rows.append(item)

    for oc in (own_checks or []):
        if hasattr(oc, 'to_dict'):
            oc = oc.to_dict()
        val = _format_currency_val(oc.get('TUTAR'))
        dov = _get_doviz_str(oc.get('DOVIZ_TIPI'))
        hesap = str(oc.get('CH_UNVANI') or oc.get('CARI_UNVAN') or 'Kendi Çekimiz')
        acik = f"Çek Ödemesi - {oc.get('BANKA') or ''} (Vade: {str(oc.get('VADE') or '')[:10]})"
        sec_ode1_rows.append((hesap, acik, val if dov == 'TL' else 0.0, val if dov == 'USD' else 0.0, val if dov == 'EUR' else 0.0))

    for cr in (credits or []):
        if hasattr(cr, 'to_dict'):
            cr = cr.to_dict()
        val = _format_currency_val(cr.get('TUTAR'))
        dov = _get_doviz_str(cr.get('DOVIZ_TIPI'))
        hesap = str(cr.get('BANKA_KREDI') or 'Banka Kredisi')
        acik = "Kredi Taksidi"
        sec_ode1_rows.append((hesap, acik, val if dov == 'TL' else 0.0, val if dov == 'USD' else 0.0, val if dov == 'EUR' else 0.0))

    for cp in (custom_payments or []):
        if hasattr(cp, 'to_dict'):
            cp = cp.to_dict()
        is_paid = (cp.get('odendi') == 1 or cp.get('odendi') is True)
        if not is_paid:
            continue
        is_cc = cp.get('credit_card_pay') or str(cp.get('category') or cp.get('kategori') or '').strip().lower() in ('kredi kartı', 'kredi karti', 'kredi kartı ödemesi')
        val = _format_currency_val(cp.get('tutar') if 'tutar' in cp else cp.get('amount'))
        dov = _get_doviz_str(cp.get('doviz'))
        hesap = f"Manuel Ödeme ({str(cp.get('kategori') or cp.get('category') or '').upper()})"
        acik = str(cp.get('aciklama') or cp.get('explanation') or '-')
        item = (hesap, acik, val if dov == 'TL' else 0.0, val if dov == 'USD' else 0.0, val if dov == 'EUR' else 0.0)
        if is_cc:
            sec_cc_rows.append(item)
        else:
            sec_ode1_rows.append(item)

    right_sections.append({'title': "1. BANKA ÖDEMELERİ (HAVALE / EFT / VİRMAN)", 'rows': sec_ode1_rows})
    right_sections.append({'title': "2. KREDİ KARTI İŞLEMLERİ", 'rows': sec_cc_rows})
    right_sections.append({'title': "3. KASA ÖDEMELERİ", 'rows': sec_kasa_rows})

    # Render side by side
    left_cur_row = current_row
    right_cur_row = current_row

    # Render left sections
    for sec in left_sections:
        ws.merge_cells(start_row=left_cur_row, start_column=1, end_row=left_cur_row, end_column=5)
        c = ws.cell(row=left_cur_row, column=1, value=sec['title'])
        c.font = subhdr_tah_font
        c.fill = subhdr_tah_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[left_cur_row].height = 20
        left_cur_row += 1

        write_headers(left_cur_row, 1)
        left_cur_row += 1

        tot_tl, tot_usd, tot_eur = 0.0, 0.0, 0.0
        if sec['rows']:
            for item in sec['rows']:
                hesap, acik, tl, usd, eur = item
                tot_tl += tl
                tot_usd += usd
                tot_eur += eur
                
                c1 = ws.cell(row=left_cur_row, column=1, value=hesap)
                c2 = ws.cell(row=left_cur_row, column=2, value=acik)
                c3 = ws.cell(row=left_cur_row, column=3, value=tl if tl > 0 else None)
                c4 = ws.cell(row=left_cur_row, column=4, value=usd if usd > 0 else None)
                c5 = ws.cell(row=left_cur_row, column=5, value=eur if eur > 0 else None)
                
                for idx, cell in enumerate([c1, c2, c3, c4, c5]):
                    cell.font = data_font
                    cell.border = cell_border
                    if idx >= 2:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                
                c3.number_format = num_fmt_tl
                c4.number_format = num_fmt_usd
                c5.number_format = num_fmt_eur
                ws.row_dimensions[left_cur_row].height = 18
                left_cur_row += 1
        else:
            ws.merge_cells(start_row=left_cur_row, start_column=1, end_row=left_cur_row, end_column=5)
            c = ws.cell(row=left_cur_row, column=1, value="Kayıt bulunamadı.")
            c.font = data_sub_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = cell_border
            ws.row_dimensions[left_cur_row].height = 18
            left_cur_row += 1

        # Section Total
        ws.merge_cells(start_row=left_cur_row, start_column=1, end_row=left_cur_row, end_column=2)
        c_tot = ws.cell(row=left_cur_row, column=1, value="Grup Toplamı")
        c_tot.font = total_font
        c_tot.fill = total_fill
        c_tot.border = cell_border
        
        c3 = ws.cell(row=left_cur_row, column=3, value=tot_tl if tot_tl > 0 else 0.0)
        c4 = ws.cell(row=left_cur_row, column=4, value=tot_usd if tot_usd > 0 else 0.0)
        c5 = ws.cell(row=left_cur_row, column=5, value=tot_eur if tot_eur > 0 else 0.0)
        
        for idx, cell in enumerate([c3, c4, c5]):
            cell.font = total_font
            cell.fill = total_fill
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        c3.number_format = num_fmt_tl
        c4.number_format = num_fmt_usd
        c5.number_format = num_fmt_eur
        ws.row_dimensions[left_cur_row].height = 19
        left_cur_row += 2

    # Render right sections
    for sec in right_sections:
        ws.merge_cells(start_row=right_cur_row, start_column=7, end_row=right_cur_row, end_column=11)
        c = ws.cell(row=right_cur_row, column=7, value=sec['title'])
        c.font = subhdr_ode_font
        c.fill = subhdr_ode_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[right_cur_row].height = 20
        right_cur_row += 1

        write_headers(right_cur_row, 7)
        right_cur_row += 1

        tot_tl, tot_usd, tot_eur = 0.0, 0.0, 0.0
        if sec['rows']:
            for item in sec['rows']:
                hesap, acik, tl, usd, eur = item
                tot_tl += tl
                tot_usd += usd
                tot_eur += eur
                
                c1 = ws.cell(row=right_cur_row, column=7, value=hesap)
                c2 = ws.cell(row=right_cur_row, column=8, value=acik)
                c3 = ws.cell(row=right_cur_row, column=9, value=tl if tl > 0 else None)
                c4 = ws.cell(row=right_cur_row, column=10, value=usd if usd > 0 else None)
                c5 = ws.cell(row=right_cur_row, column=11, value=eur if eur > 0 else None)
                
                for idx, cell in enumerate([c1, c2, c3, c4, c5]):
                    cell.font = data_font
                    cell.border = cell_border
                    if idx >= 2:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                
                c3.number_format = num_fmt_tl
                c4.number_format = num_fmt_usd
                c5.number_format = num_fmt_eur
                ws.row_dimensions[right_cur_row].height = 18
                right_cur_row += 1
        else:
            ws.merge_cells(start_row=right_cur_row, start_column=7, end_row=right_cur_row, end_column=11)
            c = ws.cell(row=right_cur_row, column=7, value="Kayıt bulunamadı.")
            c.font = data_sub_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = cell_border
            ws.row_dimensions[right_cur_row].height = 18
            right_cur_row += 1

        # Section Total
        ws.merge_cells(start_row=right_cur_row, start_column=7, end_row=right_cur_row, end_column=8)
        c_tot = ws.cell(row=right_cur_row, column=7, value="Grup Toplamı")
        c_tot.font = total_font
        c_tot.fill = total_fill
        c_tot.border = cell_border
        
        c3 = ws.cell(row=right_cur_row, column=9, value=tot_tl if tot_tl > 0 else 0.0)
        c4 = ws.cell(row=right_cur_row, column=10, value=tot_usd if tot_usd > 0 else 0.0)
        c5 = ws.cell(row=right_cur_row, column=11, value=tot_eur if tot_eur > 0 else 0.0)
        
        for idx, cell in enumerate([c3, c4, c5]):
            cell.font = total_font
            cell.fill = total_fill
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
        c3.number_format = num_fmt_tl
        c4.number_format = num_fmt_usd
        c5.number_format = num_fmt_eur
        ws.row_dimensions[right_cur_row].height = 19
        right_cur_row += 2

    # Advance current_row to max of both sides
    current_row = max(left_cur_row, right_cur_row) + 1

    # --- BOTTOM TWO SECTIONS: LEFT (BANKA BAKİYELERİ), RIGHT (ÖDEME PLANI - SONRAKİ İŞ GÜNÜ) ---
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    c_bb = ws.cell(row=current_row, column=1, value="BANKA BAKİYE DURUM RAPORU")
    c_bb.font = section_font
    c_bb.fill = section_blue_fill
    c_bb.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=11)
    plan_title = f"ÖDEME PLANI — BİR SONRAKİ İŞ GÜNÜ ({next_business_date_formatted})" if next_business_date_formatted else "ÖDEME PLANI — BİR SONRAKİ İŞ GÜNÜ"
    c_op = ws.cell(row=current_row, column=7, value=plan_title)
    c_op.font = section_font
    c_op.fill = section_amber_fill
    c_op.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    bot_left_cur_row = current_row
    bot_right_cur_row = current_row

    # Bank Balances (Left)
    # Headers
    bb_headers = ["BANKA ADI", "HESAP / AÇIKLAMA", "TL", "USD", "EUR"]
    for i, h in enumerate(bb_headers):
        c = ws.cell(row=bot_left_cur_row, column=1 + i, value=h)
        c.font = th_font
        c.fill = th_fill
        c.border = cell_border
        c.alignment = Alignment(horizontal="right" if i >= 2 else "left", vertical="center")
    ws.row_dimensions[bot_left_cur_row].height = 20
    bot_left_cur_row += 1

    # Group bank records by bank name
    grouped_banks = {}
    for r in (bank_records or []):
        if hasattr(r, 'to_dict'):
            r = r.to_dict()
        bname = str(r.get('BANKA_ADI') or r.get('banka_adi') or '-').strip()
        hname = str(r.get('HESAP_ADI') or r.get('hesap_adi') or '').strip()
        clean_bname = bname.replace(' A.Ş.', '').replace(' T.A.Ş.', '').replace(' T.A.O.', '').replace(' A.O.', '').strip()
        if not clean_bname:
            clean_bname = 'DİĞER BANKALAR'
            
        bakiye = _format_currency_val(r.get('BAKIYE') if 'BAKIYE' in r else (r.get('BAKİYE') if 'BAKİYE' in r else r.get('bakiye', 0.0)))
        dov = _get_doviz_str(r.get('DOVIZ_TIPI') if 'DOVIZ_TIPI' in r else (r.get('PARABİRİMİ') if 'PARABİRİMİ' in r else r.get('doviz_tipi', '0')))
        
        if clean_bname not in grouped_banks:
            grouped_banks[clean_bname] = {'tl': 0.0, 'usd': 0.0, 'eur': 0.0, 'accounts': []}
        
        if dov == 'TL':
            grouped_banks[clean_bname]['tl'] += bakiye
        elif dov == 'USD':
            grouped_banks[clean_bname]['usd'] += bakiye
        elif dov == 'EUR':
            grouped_banks[clean_bname]['eur'] += bakiye
            
        if hname and hname not in grouped_banks[clean_bname]['accounts']:
            grouped_banks[clean_bname]['accounts'].append(hname)

    tot_bb_tl = 0.0
    tot_bb_usd = 0.0
    tot_bb_eur = 0.0

    if grouped_banks:
        for bname, data in sorted(grouped_banks.items(), key=lambda x: x[1]['tl'], reverse=True):
            tl = data['tl']
            usd = data['usd']
            eur = data['eur']
            tot_bb_tl += tl
            tot_bb_usd += usd
            tot_bb_eur += eur
            
            c1 = ws.cell(row=bot_left_cur_row, column=1, value=bname)
            c2 = ws.cell(row=bot_left_cur_row, column=2, value=", ".join(data['accounts'][:2]) if data['accounts'] else "-")
            c3 = ws.cell(row=bot_left_cur_row, column=3, value=tl if tl != 0 else None)
            c4 = ws.cell(row=bot_left_cur_row, column=4, value=usd if usd != 0 else None)
            c5 = ws.cell(row=bot_left_cur_row, column=5, value=eur if eur != 0 else None)
            
            for idx, cell in enumerate([c1, c2, c3, c4, c5]):
                cell.font = data_font
                cell.border = cell_border
                if idx >= 2:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            c3.number_format = num_fmt_tl
            c4.number_format = num_fmt_usd
            c5.number_format = num_fmt_eur
            ws.row_dimensions[bot_left_cur_row].height = 18
            bot_left_cur_row += 1
    else:
        ws.merge_cells(start_row=bot_left_cur_row, start_column=1, end_row=bot_left_cur_row, end_column=5)
        c = ws.cell(row=bot_left_cur_row, column=1, value="Banka bakiyesi bilgisi bulunamadı.")
        c.font = data_sub_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border
        ws.row_dimensions[bot_left_cur_row].height = 18
        bot_left_cur_row += 1

    # Bank Balance Total
    ws.merge_cells(start_row=bot_left_cur_row, start_column=1, end_row=bot_left_cur_row, end_column=2)
    c_tot = ws.cell(row=bot_left_cur_row, column=1, value="BANKA BAKİYESİ TOPLAMI")
    c_tot.font = total_font
    c_tot.fill = subhdr_tah_fill
    c_tot.border = thick_top_border
    
    final_bb_tl = total_bank_balance_tl if total_bank_balance_tl else tot_bb_tl
    final_bb_usd = total_bank_balance_usd if total_bank_balance_usd else tot_bb_usd
    final_bb_eur = total_bank_balance_eur if total_bank_balance_eur else tot_bb_eur

    c3 = ws.cell(row=bot_left_cur_row, column=3, value=final_bb_tl)
    c4 = ws.cell(row=bot_left_cur_row, column=4, value=final_bb_usd)
    c5 = ws.cell(row=bot_left_cur_row, column=5, value=final_bb_eur)
    
    for idx, cell in enumerate([c3, c4, c5]):
        cell.font = total_font
        cell.fill = subhdr_tah_fill
        cell.border = thick_top_border
        cell.alignment = Alignment(horizontal="right", vertical="center")
        
    c3.number_format = num_fmt_tl
    c4.number_format = num_fmt_usd
    c5.number_format = num_fmt_eur
    ws.row_dimensions[bot_left_cur_row].height = 20
    bot_left_cur_row += 1

    # Next Day Payments (Right)
    plan_headers = ["ÖDEME TÜRÜ", "BANKA / KURUM", "AÇIKLAMA", "TUTAR", "PARA BİRİMİ"]
    for i, h in enumerate(plan_headers):
        c = ws.cell(row=bot_right_cur_row, column=7 + i, value=h)
        c.font = th_font
        c.fill = th_fill
        c.border = cell_border
        c.alignment = Alignment(horizontal="right" if i == 3 else ("center" if i == 4 else "left"), vertical="center")
    ws.row_dimensions[bot_right_cur_row].height = 20
    bot_right_cur_row += 1

    tot_plan_tl, tot_plan_usd, tot_plan_eur = 0.0, 0.0, 0.0
    if next_day_payments:
        for p in next_day_payments:
            if hasattr(p, 'to_dict'):
                p = p.to_dict()
            tur = str(p.get('tur') or '-')
            banka = str(p.get('banka') or '-')
            acik = str(p.get('aciklama') or '-')
            val = _format_currency_val(p.get('tutar'))
            dov = _get_doviz_str(p.get('doviz'))
            
            if dov == 'TL':
                tot_plan_tl += val
            elif dov == 'USD':
                tot_plan_usd += val
            elif dov == 'EUR':
                tot_plan_eur += val
                
            c1 = ws.cell(row=bot_right_cur_row, column=7, value=tur)
            c2 = ws.cell(row=bot_right_cur_row, column=8, value=banka)
            c3 = ws.cell(row=bot_right_cur_row, column=9, value=acik)
            c4 = ws.cell(row=bot_right_cur_row, column=10, value=val)
            c5 = ws.cell(row=bot_right_cur_row, column=11, value=dov)
            
            for idx, cell in enumerate([c1, c2, c3, c4, c5]):
                cell.font = data_font
                cell.border = cell_border
                if idx == 3:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif idx == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
            if dov == 'TL':
                c4.number_format = num_fmt_tl
            elif dov == 'USD':
                c4.number_format = num_fmt_usd
            elif dov == 'EUR':
                c4.number_format = num_fmt_eur
            else:
                c4.number_format = '#,##0.00'
                
            ws.row_dimensions[bot_right_cur_row].height = 18
            bot_right_cur_row += 1
    else:
        ws.merge_cells(start_row=bot_right_cur_row, start_column=7, end_row=bot_right_cur_row, end_column=11)
        c = ws.cell(row=bot_right_cur_row, column=7, value="Bir sonraki iş günü için vadesi gelen ödeme bulunmamaktadır.")
        c.font = data_sub_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border
        ws.row_dimensions[bot_right_cur_row].height = 18
        bot_right_cur_row += 1

    # Auto-fit Column Widths
    col_widths = {
        1: 28,  # Hesap / Kasa
        2: 32,  # Açıklama
        3: 16,  # TL
        4: 16,  # USD
        5: 16,  # EUR
        6: 4,   # Gap
        7: 24,  # Sağ Hesap / Tür
        8: 24,  # Banka / Açıklama
        9: 28,  # Açıklama
        10: 16, # Tutar
        11: 12  # Para Birimi
    }
    
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
